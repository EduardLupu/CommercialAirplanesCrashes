from bs4 import BeautifulSoup
from openpyxl import Workbook
import requests
from functions import *

if __name__ == '__main__':
    response = requests.get(URL)
    if response.status_code == 404:
        print("No response from wiki! Use local database!")
        exit(0)
    wiki_text = response.text
    parse = BeautifulSoup(wiki_text, 'lxml')
    accidents = parse.find_all('li', {'class': None, 'id': None})

    excel = Workbook()
    sheet = excel.create_sheet(title="CommercialAirplanesCrashesData")
    index = 1

    for accident in accidents:
        year = accident.find_previous('span', {'class': 'mw-headline'}).text + '|\n'
        accident_info = accident.text.replace('–', year)
        accident_date = accident_info[:accident_info.find('|')]
        if check_month(accident_info) is False:
            break
        if len(accident_date) <= 17:
            dates.append(accident_date)
            number_of_accidents += 1
            sheet.cell(row=index, column=1, value=accident_date)
            sheet.cell(row=index, column=2, value=accident_info[accident_info.find('|') + 1:])
            sheet.merge_cells(start_row=index, start_column=2, end_row=index, end_column=42)
            index += 1

    add_months()
    del excel['Sheet']
    excel.save(SAVED_INFORMATION)
    days_of_the_week()
    seasons()

    print(f"\n\nTotal numbers of commercial airplanes crashes from 1919 to 2022: {number_of_accidents}\n")
    print("Commercial Airplanes Accidents by seasons: ")
    print(f"Spring: {seasons_value[0]}    Percentage: {percentage(seasons_value[0], number_of_accidents)}")
    print(f"Summer: {seasons_value[1]}    Percentage: {percentage(seasons_value[1], number_of_accidents)}")
    print(f"Autumn: {seasons_value[2]}    Percentage: {percentage(seasons_value[2], number_of_accidents)}")
    print(f"Winter: {seasons_value[3]}    Percentage: {percentage(seasons_value[3], number_of_accidents)}")

    print("\nCommercial Airplanes Accidents by day of the week:")
    print(f"Monday:    {days_value[1]}    Percentage: {percentage(days_value[1], number_of_accidents)}")
    print(f"Tuesday:   {days_value[2]}    Percentage: {percentage(days_value[2], number_of_accidents)}")
    print(f"Wednesday: {days_value[3]}    Percentage: {percentage(days_value[3], number_of_accidents)}")
    print(f"Thursday:  {days_value[4]}    Percentage: {percentage(days_value[4], number_of_accidents)}")
    print(f"Friday:    {days_value[5]}    Percentage: {percentage(days_value[5], number_of_accidents)}")
    print(f"Saturday:  {days_value[6]}    Percentage: {percentage(days_value[6], number_of_accidents)}")
    print(f"Sunday:    {days_value[0]}    Percentage: {percentage(days_value[0], number_of_accidents)}")
